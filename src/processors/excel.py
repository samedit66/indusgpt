import logging
import pathlib

import openpyxl

from src import types
from src.chat import info_extractor
from src.persistence import models
from src.processors import utils


class ExcelProcessor:
    """Processor that writes Q&A data to an xlsx file."""

    def __init__(
        self,
        sheet_title: str,
        output_path: str | None = None,
    ) -> None:
        self.sheet_title = sheet_title
        self.output_path = output_path

    async def __call__(self, user_id: int, qa_pairs: list[types.QaPair]):
        # Fetch user details
        user = await models.User.filter(id=user_id).first()
        user_name = user.name
        tg = user.url

        # Extract and flatten the information
        user_info = await info_extractor.extract_info(qa_pairs)
        row = utils.flatten_user_info(user_name, tg, user_info)

        # Load existing workbook or create a new one
        if self.output_path:
            output_path = pathlib.Path(self.output_path)
            if not output_path.exists():
                output_path.touch()

            wb = openpyxl.load_workbook(self.output_path)
        else:
            wb = openpyxl.Workbook()

        # Create or get the sheet
        if self.sheet_title in wb.sheetnames:
            sheet = wb[self.sheet_title]
        else:
            sheet = wb.create_sheet(title=self.sheet_title)

        # Write header if sheet is new/empty
        if (
            sheet.max_row == 1
            and sheet.max_column == 1
            and sheet.cell(1, 1).value is None
        ):
            headers = list(row.keys())
            sheet.append(headers)

        # Append the data row in order of headers
        headers = [cell.value for cell in sheet[1]]
        sheet.append([row.get(h, "") for h in headers])

        if not self.output_path:
            return wb

        wb.save(output_path)
        logging.info("Written Excel row for user %s to %s", user_name, self.output_path)
